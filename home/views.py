import math
import random

import numpy as np
import pandas as pd
import scipy.stats
from django.shortcuts import render
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font


def home(request):
    return render(request,'home.html')


def sim_process_row(row, ci, simulation_duration_type,deviation):
    if pd.isna(row['ItemNumber']):
        return row

    if ci == 1:
        cs = row['Current Stock']
        prev_stock = 0 if pd.isna(cs) else cs
    else:
        prev_stock = row[simulation_duration_type + str(ci - 1)]

    current_index = simulation_duration_type + str(ci)


    if simulation_duration_type == 'M':
        if float(deviation) == float(0):
            average_consumption = 0.00 if pd.isna(row['Avrg. Consumption/Mnth']) else float(
                row['Avrg. Consumption/Mnth'])
        else:
            average_consumption = 0.00 if pd.isna(row['Avrg. Consumption/Mnth']) else float(
                row['Avrg. Consumption/Mnth'])
            limit = float(average_consumption * (deviation / 100))
            average_consumption = random.uniform(average_consumption - limit, average_consumption + limit)
    elif simulation_duration_type == 'D':
        if float(deviation) == float(0):
            average_consumption = 0.00 if pd.isna(row['Avrg. Consumption/Mnth']) else float(
                row['Avrg. Consumption/Mnth'])
            average_consumption = average_consumption / 30.416
        else:
            average_consumption = 0.00 if pd.isna(row['Avrg. Consumption/Mnth']) else float(
                row['Avrg. Consumption/Mnth'])

            average_consumption = average_consumption / 30.416
            limit = float(average_consumption * (deviation / 100))
            average_consumption = random.uniform(average_consumption - limit, average_consumption + limit)
    print(prev_stock,average_consumption)
    diff = float(prev_stock) - average_consumption

    if pd.isna(row['Prev_Order']):
        row[current_index] = diff

        if row[current_index] <= row['MIN (Reorder Point)']:
            order = row['MAX (Order Up To)'] - diff
            if simulation_duration_type == 'M':

                if row['Rounded Lead Time'] < 1:
                    row[current_index] = diff + order
                    row['Prev_Order'] = pd.NA
                else:
                    row['Prev_Order'] = str(int(current_index.lstrip(simulation_duration_type)) + math.floor(
                        row['Rounded Lead Time'])) + '+' + str(order)
            else:
                row['Prev_Order'] = str(
                    int(current_index.lstrip(simulation_duration_type)) + row['Rounded Lead Time']) + '+' + str(
                    order)
    else:
        if str(int(current_index.lstrip(simulation_duration_type))) == row['Prev_Order'].split('+')[0].rstrip('.0'):
            row[current_index] = diff + float(row['Prev_Order'].split('+')[1])
            row['Prev_Order'] = pd.NA

            if row[current_index] <= row['MIN (Reorder Point)']:
                order = row['MAX (Order Up To)'] - row[current_index]
                row['Prev_Order'] = str(
                    int(current_index.lstrip(simulation_duration_type)) + row['Rounded Lead Time']) + '+' + str(
                    order)
        else:
            row[current_index] = diff

    return row


def sim_save_output(df, fs_df):
    output_file = 'output/sim_output.xlsx'

    try:
        writer = pd.ExcelWriter(output_file, engine='openpyxl', mode='w+')

        df.to_excel(writer, sheet_name='Inventory Simulation', index=False)
        fs_df.to_excel(writer, sheet_name='Financial Simulation', index=False)

        writer.close()

        ExcelWorkbook = load_workbook(output_file)
        for sheet_name in ['Inventory Simulation', 'Financial Simulation']:
            ws = ExcelWorkbook[sheet_name]
            thin_border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))
            max_column = ws.max_column + 1
            max_row = ws.max_row + 1
            for col in range(1, max_column):
                ws.cell(1, col).fill = PatternFill("solid", start_color="d9e1f2")
                for row in range(1, max_row):
                    ws.cell(row, col).border = thin_border
                    ws.cell(row, col).font = Font(name='Trebuchet MS')
                    if row == max_row - 1 and sheet_name == 'Financial Simulation':
                        ws.cell(row, col).fill = PatternFill("solid", start_color="C5D9F1")

        ExcelWorkbook.save(output_file)

    except Exception as e:
        print(e)
        print('Unable To Save File')


def simulation(request):
    if request.method == 'POST':
        # Retrieve user-provided Python code from POST request
        simulation_time = int(request.POST.get('simulation_period'))
        simulation_time_type = request.POST.get('months')
        deviation = float(request.POST.get('consumption_deviation'))
        default_price = request.POST.get('price_unit')
        removeNegative = request.POST.get('removeNegative')  # Checkbox value for "Remove Negative"
        roundLeadTime = request.POST.get('roundLeadTime')  # Checkbox value for "Round Lead Time"
        browse = request.FILES.get('filedata')

        # Initialize or load your DataFrame (df) here
        try:
            if browse is not None:
                df = pd.read_excel(browse)
            else:
                df = pd.DataFrame()  # Initialize an empty DataFrame if no file is uploaded
                return
        except Exception as e:
            # Handle any potential errors when reading the file
            print("Error reading the file:", str(e))
            df = pd.DataFrame()  # Initialize an empty DataFrame as a fallback
            return

        df = df.dropna(how='all')
        df = df.fillna(0)
        df = df.replace({0: pd.NA, '0': pd.NA, '': pd.NA, 'None': pd.NA})

        simulation_duration_type = 'M'
        if simulation_time_type == 'months':
            df['Lead Time (months)'] = df['Lead Time (days)'].astype(float) / 30.416
            if roundLeadTime.isChecked():
                df['Rounded Lead Time'] = df['Lead Time (months)'].replace({pd.NA: 0}).round().astype(int)
            else:
                df['Rounded Lead Time'] = df['Lead Time (months)'].replace({pd.NA: 0}).astype(float)
            simulation_duration_type = 'M'
        elif simulation_time_type == 'days':
            simulation_duration_type = 'D'
            df['Rounded Lead Time'] = df['Lead Time (days)'].replace({pd.NA: 0}).astype(float)

        df['Prev_Order'] = pd.NA

        for rep in range(simulation_time):
            df['{}{}'.format(simulation_duration_type, rep + 1)] = ''
            for ci in range(1, len(df) + 1):
                row = df.loc[ci - 1]
                new_row = sim_process_row(row, ci, simulation_duration_type,deviation)
                df.loc[ci - 1] = new_row

        df.drop(columns=['Prev_Order'], inplace=True)

        fs_df = df[['ItemNumber', 'Description', 'Price/Unit']].copy()

        for column in range(simulation_time):
            colm_title = '{}{}'.format(simulation_duration_type, column + 1)
            price_column = df['Price/Unit']
            price_column.replace({pd.NA: default_price}, inplace=True)
            amount_column = df[colm_title] * price_column
            fs_df[colm_title] = amount_column

        def remove_neg(x):
            if 'M' in x.name:
                return x.clip(lower=0)
            else:
                return x

        if removeNegative.isChecked():
            fs_df = fs_df.apply(remove_neg)

        fs_df.loc['Total'] = fs_df.sum(axis=0)
        fs_df.iloc[-1, 1] = ''
        fs_df.iloc[-1, 0] = 'Total'
        print('im here')
        sim_save_output(df, fs_df)

        return render(request, 'simulate.html', {'result': 'Simulation completed and saved successfully.'})

    # Handle GET request here
    return render(request, 'simulate.html')


# optimizer functions
def opt_nonzerostd(row):
    row = opt_filtered_row(row)
    row = row.replace({0: pd.NA, '0': pd.NA, '': pd.NA, 'None': pd.NA})
    row = pd.Series(row.dropna().reset_index(drop=True))

    return row.std(ddof=0)

def opt_filtered_row(row):
    first_index = row.replace({0: pd.NA, '0': pd.NA, '': pd.NA, 'None': pd.NA}).first_valid_index()
    row = row[first_index:]
    row = row.fillna(0)
    return pd.Series(row.replace({pd.NA: 0}))

def opt_length(row):
    return len(opt_filtered_row(row))

def opt_save_output(df):
    fileName='output/opt_output.xlsx'
    df.to_excel(fileName, index=False)

    ExcelWorkbook = load_workbook(fileName)

    # ExcelWorkbook = load_workbook(self.ui.ExcelPath.text())
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    ws = ExcelWorkbook.active
    max_column = ws.max_column + 1
    max_row = ws.max_row + 1
    for col in range(1, max_column):
        ws.cell(1, col).fill = PatternFill("solid", start_color="d9e1f2")
        for row in range(1, max_row):
            ws.cell(row, col).border = thin_border

    ExcelWorkbook.save(fileName)

def optimizer(request):
    if request.method == 'POST':
        lead_time = request.POST.get('lead_time')
        service_level = request.POST.get('service_level')
        review_period = request.POST.get('review_period')
        benchmark = request.POST.get('benchmark')
        browse = request.FILES.get('filedata')
        print("xxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxx", lead_time)

        LT_Adjustment = 0 if lead_time in [None, 'None', '', ' '] else int(lead_time)
        Alpha = 0 if service_level in [None, 'None', '', ' '] else float(service_level)
        ReviewPeriod = 0 if review_period in [None, 'None', '', ' '] else float(review_period) / 30.41666667

        # Initialize or load your DataFrame (df) here
        try:
            if browse is not None:
                df = pd.read_excel(browse)
            else:
                df = pd.DataFrame()  # Initialize an empty DataFrame if no file is uploaded
                return
        except Exception as e:
            # Handle any potential errors when reading the file
            print("Error reading the file:", str(e))
            df = pd.DataFrame()  # Initialize an empty DataFrame as a fallback
            return


        print('input file loaded')



        start_col = df.columns.get_loc('Description')

        end_col = len(df.columns) - 3
        current_min_col = df.iloc[:, end_col + 1]

        if not df.empty:
            # Perform calculations on df
            df['Total'] = df.iloc[:, start_col + 1:end_col].apply(opt_filtered_row, axis=1).sum(axis=1)
            df['LifeTime'] = df.iloc[:, start_col + 1:end_col].apply(opt_length, axis=1)
            df['Min Order'] = df.iloc[:, start_col + 1:end_col].apply(opt_filtered_row, axis=1).replace(
                {'0': pd.NA, 0: pd.NA}).min(axis=1)
            df['Max Order'] = df.iloc[:, start_col + 1:end_col].apply(opt_filtered_row, axis=1).max(axis=1)
            df['Adjusted LT (days)'] = df.iloc[:, end_col] + LT_Adjustment
            df['Adjusted LT (month)'] = df['Adjusted LT (days)'] / 30.41666667
            df['Average Demand'] = df['Total'] / df['LifeTime']
            df['Demand over LT'] = df['Average Demand'] * df['Adjusted LT (month)']
            df['Std. Deviation of Demand'] = df.iloc[:, start_col + 1:end_col].apply(opt_nonzerostd, 1)
            df['Std. Dev. of Demand over LT'] = df['Std. Deviation of Demand'] * np.sqrt(
                df['Adjusted LT (month)'].astype(float))
            df['Safety Stock'] = scipy.stats.norm.ppf(Alpha, 0, 1) * df['Std. Dev. of Demand over LT']
            df['Min'] = scipy.stats.norm.ppf(Alpha, 0, 1) * df['Std. Dev. of Demand over LT'] + df['Demand over LT']

            if benchmark == 'No BenchMark':
                df['Adjusted MIN'] = df['Min']
            elif benchmark == 'Current Min':
                df['Adjusted MIN'] = np.where(df['Min'] < current_min_col, current_min_col, df['Min'])
            elif benchmark == 'Max Order':
                df['Adjusted MIN'] = np.where(df['Min'] < df['Max Order'], df['Max Order'], df['Min'])
            elif benchmark == 'Both (Current Min + Max Order)':
                max_result = np.where(current_min_col < df['Max Order'], df['Max Order'], current_min_col)
                df['Adjusted MIN'] = np.where(df['Min'] < max_result, max_result, df['Min'])
            elif benchmark == 'Slow Moving Items':
                df['Adjusted MIN'] = np.ceil(
                    np.where(df['Demand over LT'] > 10, df['Demand over LT'] + df['Average Demand'],
                             df['Demand over LT'] + 1)).astype(int)
                df['MAX'] = np.ceil(df['Adjusted MIN'] + df['Demand over LT']).astype(int)

            elif benchmark == 'Compare All':
                df['Adjusted MIN-No BenchMark'] = df['Min']
                df['Adjusted MIN-Current Min'] = np.where(df['Min'] < current_min_col, current_min_col, df['Min'])
                df['Adjusted MIN-Max Order'] = np.where(df['Min'] < df['Max Order'], df['Max Order'], df['Min'])
                max_result = np.where(current_min_col < df['Max Order'], df['Max Order'], current_min_col)
                df['Adjusted MIN-Both (Current Min + Max Order)'] = np.where(df['Min'] < max_result, max_result, df['Min'])
                df['Adjusted MIN-Slow Moving Items'] = np.ceil(
                    np.where(df['Demand over LT'] > 10, df['Demand over LT'] + df['Average Demand'],
                             df['Demand over LT'] + 1)).astype(int)

                df['MAX-No BenchMark'] = (df['Average Demand'] * ReviewPeriod) + df['Adjusted MIN-No BenchMark']
                df['MAX-Current Min'] = (df['Average Demand'] * ReviewPeriod) + df['Adjusted MIN-Current Min']
                df['MAX-Max Order'] = (df['Average Demand'] * ReviewPeriod) + df['Adjusted MIN-Max Order']
                df['MAX-Both (Current Min + Max Order)'] = (df['Average Demand'] * ReviewPeriod) + df['Adjusted MIN-Both']
                df['MAX-Slow Moving Items'] = np.ceil(df['Adjusted MIN-Slow Moving Items'] + df['Demand over LT']).astype(int)

            if benchmark != 'Slow Moving Items' and benchmark != 'Compare All':
                df['MAX'] = (df['Average Demand'] * ReviewPeriod) + df['Adjusted MIN']

            if benchmark == 'Compare All':
                df['Cycle Service Level-No BenchMark'] = scipy.stats.norm.cdf(
                    (df['Adjusted MIN-No BenchMark'].astype(float) + 0.0056).round(decimals=4),
                    df['Demand over LT'].astype(float).round(decimals=3),
                    df['Std. Dev. of Demand over LT'].replace({0: 0.001}).astype(float).round(decimals=4))

                df['Cycle Service Level-Current Min'] = scipy.stats.norm.cdf(
                    (df['Adjusted MIN-Current Min'].astype(float) + 0.0056).round(decimals=4),
                    df['Demand over LT'].astype(float).round(decimals=3),
                    df['Std. Dev. of Demand over LT'].replace({0: 0.001}).astype(float).round(decimals=4))

                df['Cycle Service Level-Max Order'] = scipy.stats.norm.cdf(
                    (df['Adjusted MIN-Max Order'].astype(float) + 0.0056).round(decimals=4),
                    df['Demand over LT'].astype(float).round(decimals=3),
                    df['Std. Dev. of Demand over LT'].replace({0: 0.001}).astype(float).round(decimals=4))

                df['Cycle Service Level-Both (Current Min + Max Order)'] = scipy.stats.norm.cdf(
                    (df['Adjusted MIN-Both (Current Min + Max Order)'].astype(float) + 0.0056).round(decimals=4),
                    df['Demand over LT'].astype(float).round(decimals=3),
                    df['Std. Dev. of Demand over LT'].replace({0: 0.001}).astype(float).round(decimals=4))

                df['Cycle Service Level-Slow Moving Items'] = scipy.stats.norm.cdf(
                    (df['Adjusted MIN-Slow Moving Items'].astype(float) + 0.0056).round(decimals=4),
                    df['Demand over LT'].astype(float).round(decimals=3),
                    df['Std. Dev. of Demand over LT'].replace({0: 0.001}).astype(float).round(decimals=4))
            else:
                df['Cycle Service Level'] = scipy.stats.norm.cdf(
                    (df['Adjusted MIN'].astype(float) + 0.0056).round(decimals=4),
                    df['Demand over LT'].astype(float).round(decimals=3),
                    df['Std. Dev. of Demand over LT'].replace({0: 0.001}).astype(float).round(decimals=4))

            df['Adjusted LT (month)'] = df['Adjusted LT (month)'].astype(float).round(1)
            df['Average Demand'] = df['Average Demand'].astype(float).round(1)
            df['Demand over LT'] = df['Demand over LT'].astype(float).round(1)
            df['Std. Deviation of Demand'] = df['Std. Deviation of Demand'].astype(float).round(1)
            df['Std. Dev. of Demand over LT'] = df['Std. Dev. of Demand over LT'].astype(float).round(1)
            df['Safety Stock'] = df['Safety Stock'].astype(float).round(1)
            df['Min'] = df['Min'].astype(float).round(1)

            if benchmark == 'Compare All':
                for item in ['No BenchMark', 'Current Min', 'Max Order', 'Both (Current Min + Max Order)', 'Slow Moving Items']:
                    df['Adjusted MIN-{}'.format(item)] = df['Adjusted MIN-{}'.format(item)].astype(float).round(1)
                    df['MAX-{}'.format(item)] = df['MAX-{}'.format(item)].astype(float).round(1)
                    df['Cycle Service Level-{}'.format(item)] = df['Cycle Service Level-{}'.format(item)].astype(
                        float).round(5)
            else:
                df['Adjusted MIN'] = df['Adjusted MIN'].astype(float).round(1)
                df['MAX'] = df['MAX'].astype(float).round(1)
                df['Cycle Service Level'] = df['Cycle Service Level'].astype(float).round(5)

            # Prepare data for rendering
            context = {
                'result_data': {
                    'LT_Adjustment': LT_Adjustment,
                    'Alpha': Alpha,
                    'ReviewPeriod': ReviewPeriod,
                    'benchmark': benchmark,
                    # Include other calculated values here
                },
                'df1_results': df.to_dict(orient='records')
            }
            opt_save_output(df)
            return render(request, 'index.html', context)
        else:
            # Handle the case where the DataFrame is empty or couldn't be loaded
            return render(request, 'index.html', {'error_message': 'Error loading or processing the file'})
    else:
        return render(request, 'index.html')
