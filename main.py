
import math
import os
import random

import numpy as np
import pandas as pd
import scipy.stats
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font


class MyWindow:
    # ignore for now
    def start_both_processes(self):
        if self.ui.CTI_ExcelPath.text() is None or self.ui.CTS_ExcelPath_1.text() is None:
            return
        try:
            df1 = pd.read_excel(self.ui.CTI_ExcelPath.text())
            df2 = pd.read_excel(self.ui.CTS_ExcelPath_1.text())
        except Exception as e:
            print(e)
            return

        df1 = df1.dropna(how='all')
        df1 = df1.fillna(0)
        df1 = df1.replace({0: pd.NA, '0': pd.NA, '': pd.NA, 'None': pd.NA})

        start_col = df1.columns.get_loc('Description')

        end_col = len(df1.columns) - 3
        current_min_col = df1.iloc[:, end_col + 1].replace({pd.NA: 0})

        LT_Adjustment = 0 if self.ui.CTI_LT_Adjustment.text() in [None, 'None', '', ' '] else int(
            self.ui.CTI_LT_Adjustment.text())
        Alpha = 0 if self.ui.CTI_Alpha.text() in [None, 'None', '', ' '] else float(self.ui.CTI_Alpha.text())
        ReviewPeriod = 0 if self.ui.CTI_ReviewPeriod.text() in [None, 'None', '', ' '] else float(
            self.ui.CTI_ReviewPeriod.text())

        #from days to months
        ReviewPeriod = ReviewPeriod/ 30.41666667


        df1['Total'] = df1.iloc[:, start_col + 1:end_col].apply(self.opt_filtered_row, axis=1).sum(axis=1)
        df1['LifeTime'] = df1.iloc[:, start_col + 1:end_col].apply(self.opt_length, axis=1)
        df1['Min Order'] = df1.iloc[:, start_col + 1:end_col].apply(self.opt_filtered_row, axis=1).replace(
            {'0': pd.NA, 0: pd.NA}).min(axis=1)
        df1['Max Order'] = df1.iloc[:, start_col + 1:end_col].apply(self.opt_filtered_row, axis=1).max(axis=1)
        df1['Adjusted LT (days)'] = df1.iloc[:, end_col] + LT_Adjustment
        df1['Adjusted LT (month)'] = df1['Adjusted LT (days)'] / 30.41666667
        df1['Average Demand'] = df1['Total'] / df1[
            'LifeTime']  # df.iloc[:, start_col + 1:end_col].apply(self.get_filtered_row,axis=1).mean(axis=1)
        # #df.iloc[:, pd.Series.first_valid_index:end_col].mean(axis=1)
        df1['Demand over LT'] = df1['Average Demand'] * df1['Adjusted LT (month)']

        df1['Std. Deviation of Demand'] = df1.iloc[:, start_col + 1:end_col].apply(self.opt_nonzerostd,
                                                                                   1)  # .std(axis=1).replace({0:0,
        # pd.NA:0,np.nan:0})

        df1['Std. Dev. of Demand over LT'] = df1['Std. Deviation of Demand'] * np.sqrt(
            df1['Adjusted LT (month)'].astype(float))
        # std_of_demand_over_LT =
        df1['Safety Stock'] = scipy.stats.norm.ppf(Alpha, 0, 1) * df1[
            'Std. Dev. of Demand over LT']  # norm.ppf(alpha-Alpha,mu-,sigma-sd) norm.inv(0.98,0,1)
        df1['Min'] = scipy.stats.norm.ppf(Alpha, 0, 1) * df1['Std. Dev. of Demand over LT'] + df1['Demand over LT']
        # 'No BenchMark','Current Min','Max Order','Both','Slow Moving Items'
        if str(self.ui.CTI_comboBox.currentText()) == 'No BenchMark':
            df1['Adjusted MIN'] = df1['Min']
        elif str(self.ui.CTI_comboBox.currentText()) == 'Current Min':
            df1['Adjusted MIN'] = np.where(df1['Min'] < current_min_col, current_min_col, df1['Min'])
        elif str(self.ui.CTI_comboBox.currentText()) == 'Max Order':
            df1['Adjusted MIN'] = np.where(df1['Min'] < df1['Max Order'], df1['Max Order'], df1['Min'])
        elif str(self.ui.CTI_comboBox.currentText()) == 'Both':
            max_result = np.where(current_min_col < df1['Max Order'], df1['Max Order'], current_min_col)
            df1['Adjusted MIN'] = np.where(df1['Min'] < max_result, max_result, df1['Min'])
        elif str(self.ui.CTI_comboBox.currentText()) == 'Slow Moving Items':
            df1['Adjusted MIN'] = np.ceil(
                np.where(df1['Demand over LT'] > 10, df1['Demand over LT'] + df1['Average Demand'],
                         df1['Demand over LT'] + 1)).astype(int)
            df1['MAX'] = np.ceil(df1['Adjusted MIN'] + df1['Demand over LT']).astype(int)

        elif str(self.ui.CTI_comboBox.currentText()) == 'Compare All':
            df1['Adjusted MIN-No BenchMark'] = df1['Min']
            df1['Adjusted MIN-Current Min'] = np.where(df1['Min'] < current_min_col, current_min_col, df1['Min'])
            df1['Adjusted MIN-Max Order'] = np.where(df1['Min'] < df1['Max Order'], df1['Max Order'], df1['Min'])
            max_result = np.where(current_min_col < df1['Max Order'], df1['Max Order'], current_min_col)
            df1['Adjusted MIN-Both'] = np.where(df1['Min'] < max_result, max_result, df1['Min'])
            df1['Adjusted MIN-Slow Moving Items'] = np.ceil(
                np.where(df1['Demand over LT'] > 10, df1['Demand over LT'] + df1['Average Demand'],
                         df1['Demand over LT'] + 1)).astype(int)

            df1['MAX-No BenchMark'] = (df1['Average Demand'] * ReviewPeriod) + df1['Adjusted MIN-No BenchMark']
            df1['MAX-Current Min'] = (df1['Average Demand'] * ReviewPeriod) + df1['Adjusted MIN-Current Min']
            df1['MAX-Max Order'] = (df1['Average Demand'] * ReviewPeriod) + df1['Adjusted MIN-Max Order']
            df1['MAX-Both'] = (df1['Average Demand'] * ReviewPeriod) + df1['Adjusted MIN-Both']
            df1['MAX-Slow Moving Items'] = np.ceil(
                df1['Adjusted MIN-Slow Moving Items'] + df1['Demand over LT']).astype(int)
        # need addition over here
        if str(self.ui.CTI_comboBox.currentText()) != 'Slow Moving Items' and str(
                self.ui.CTI_comboBox.currentText()) != 'Compare All':
            df1['MAX'] = (df1['Average Demand'] * ReviewPeriod) + df1['Adjusted MIN']

        # std_of_demand=df['Std. Deviation of Demand'].replace({0:0.0001})
        # std_of_demand_over_LT=std_of_demand*np.sqrt(df['Adjusted LT (month)'].astype(float))
        # df['Adjusted STD DEV of Demand over LT']=df['Std. Dev. of Demand over LT'].replace({0: 0.001})
        if str(self.ui.CTI_comboBox.currentText()) == 'Compare All':
            df1['Cycle Service Level-No BenchMark'] = scipy.stats.norm.cdf(
                (df1['Adjusted MIN-No BenchMark'].astype(float) + 0.0056).round(decimals=4),
                df1['Demand over LT'].astype(float).round(decimals=3),
                df1['Std. Dev. of Demand over LT'].replace({0: 0.001}).astype(
                    float).round(decimals=4))

            df1['Cycle Service Level-Current Min'] = scipy.stats.norm.cdf(
                (df1['Adjusted MIN-Current Min'].astype(float) + 0.0056).round(decimals=4),
                df1['Demand over LT'].astype(float).round(decimals=3),
                df1['Std. Dev. of Demand over LT'].replace({0: 0.001}).astype(
                    float).round(decimals=4))

            df1['Cycle Service Level-Max Order'] = scipy.stats.norm.cdf(
                (df1['Adjusted MIN-Max Order'].astype(float) + 0.0056).round(decimals=4),
                df1['Demand over LT'].astype(float).round(decimals=3),
                df1['Std. Dev. of Demand over LT'].replace({0: 0.001}).astype(
                    float).round(decimals=4))

            df1['Cycle Service Level-Both'] = scipy.stats.norm.cdf(
                (df1['Adjusted MIN-Both'].astype(float) + 0.0056).round(decimals=4),
                df1['Demand over LT'].astype(float).round(decimals=3),
                df1['Std. Dev. of Demand over LT'].replace({0: 0.001}).astype(
                    float).round(decimals=4))

            df1['Cycle Service Level-Slow Moving Items'] = scipy.stats.norm.cdf(
                (df1['Adjusted MIN-Slow Moving Items'].astype(float) + 0.0056).round(decimals=4),
                df1['Demand over LT'].astype(float).round(decimals=3),
                df1['Std. Dev. of Demand over LT'].replace({0: 0.001}).astype(
                    float).round(decimals=4))
        else:
            df1['Cycle Service Level'] = scipy.stats.norm.cdf(
                (df1['Adjusted MIN'].astype(float) + 0.0056).round(decimals=4),
                df1['Demand over LT'].astype(float).round(decimals=3),
                df1['Std. Dev. of Demand over LT'].replace({0: 0.001}).astype(
                    float).round(decimals=4))
        # NormalDist(mu=0, sigma=1).cdf(1.96)

        # Formatiing
        df1['Adjusted LT (month)'] = df1['Adjusted LT (month)'].astype(float).round(1)
        df1['Average Demand'] = df1['Average Demand'].astype(float).round(1)
        df1['Demand over LT'] = df1['Demand over LT'].astype(float).round(1)
        df1['Std. Deviation of Demand'] = df1['Std. Deviation of Demand'].astype(float).round(1)
        df1['Std. Dev. of Demand over LT'] = df1['Std. Dev. of Demand over LT'].astype(float).round(1)
        df1['Safety Stock'] = df1['Safety Stock'].astype(float).round(1)
        df1['Min'] = df1['Min'].astype(float).round(1)
        if str(self.ui.CTI_comboBox.currentText()) == 'Compare All':
            for item in ['No BenchMark', 'Current Min', 'Max Order', 'Both', 'Slow Moving Items']:
                df1['Adjusted MIN-{}'.format(item)] = df1['Adjusted MIN-{}'.format(item)].astype(float).round(1)
                df1['MAX-{}'.format(item)] = df1['MAX-{}'.format(item)].astype(float).round(1)
                df1['Cycle Service Level-{}'.format(item)] = df1['Cycle Service Level-{}'.format(item)].astype(
                    float).round(5)
        else:
            df1['Adjusted MIN'] = df1['Adjusted MIN'].astype(float).round(1)
            df1['MAX'] = df1['MAX'].astype(float).round(1)
            df1['Cycle Service Level'] = df1['Cycle Service Level'].astype(float).round(5)

        # optimizer code finish

        simulation_time = self.ui.CTS_Simulation_period.value()

        if simulation_time == 0:
            ###
            ###
            ###
            return
        simulation_time_type = str(self.ui.CTS_simulation_duration_type.currentText()).lower().strip()
        default_price = self.ui.CTS_price.value()

        # df = pd.read_excel('Book5.xlsx')

        df2 = df2.dropna(how='all')
        df2 = df2.fillna(0)
        df2 = df2.replace({0: pd.NA, '0': pd.NA, '': pd.NA, 'None': pd.NA})

        # ...................................

        if simulation_time_type == 'months':
            df2['Lead Time (months)'] = df2['Lead Time (days)'].astype(float) / 30.416
            if self.ui.CTS_Round_LT.isChecked():
                df2['Rounded Lead Time'] = df2['Lead Time (months)'].replace({pd.NA: 0}).round().astype(int)
            else:
                df2['Rounded Lead Time'] = df2['Lead Time (months)'].replace({pd.NA: 0}).astype(float)
            simulation_duration_type = 'M'
        elif simulation_time_type == 'days':
            simulation_duration_type = 'D'
            df2['Rounded Lead Time'] = df2['Lead Time (days)'].replace({pd.NA: 0}).astype(float)

        df2['Prev_Order'] = pd.NA

        for rep in range(simulation_time):
            df2['{}{}'.format(simulation_duration_type, rep + 1)] = ''
            new_data = df2.apply(self.sim_process_row, axis=1, ci=rep + 1,
                                 simulation_duration_type=simulation_duration_type)
            # if not pd.isna(new_data['Prev_Order']):
            df2['Prev_Order'] = new_data['Prev_Order']
            df2['{}{}'.format(simulation_duration_type, rep + 1)] = new_data[
                '{}{}'.format(simulation_duration_type, rep + 1)]

        df2.drop(columns=['Prev_Order'], inplace=True)

        # calculate second sheet
        fs_df = df2[['ItemNumber', 'Description', 'Price/Unit']].copy()

        for column in range(simulation_time):
            colm_title = '{}{}'.format(simulation_duration_type, column + 1)
            price_column = df2['Price/Unit']
            price_column.replace({pd.NA: default_price}, inplace=True)
            amount_column = df2[colm_title] * price_column
            fs_df[colm_title] = amount_column

        def remove_neg(x):
            if 'M' in x.name:
                return x.clip(lower=0)
            else:
                return x

        if self.ui.CTS_Remove_Negs.isChecked():
            fs_df = fs_df.apply(remove_neg)

        fs_df.loc['Total'] = fs_df.sum(axis=0)
        fs_df.iloc[-1, 1] = ''
        fs_df.iloc[-1, 0] = 'Total'

        fileName='output'
        if fileName in [None, '']: return
        try:
            writer = pd.ExcelWriter(fileName, engine='openpyxl', mode='w+')
            # writer.book = Workbook()

            df2.to_excel(writer, sheet_name='Inventory Simulation', index=False)
            fs_df.to_excel(writer, sheet_name='Financial Simulation', index=False)
            df1.to_excel(writer, sheet_name='Inventory Optimizor', index=False)

            # writer.save()
            writer.close()

            ExcelWorkbook = load_workbook(fileName)
            # ExcelWorkbook.remove(ExcelWorkbook['Sheet'])
            for sheet_name in ['Inventory Simulation', 'Financial Simulation', 'Inventory Optimizor']:
                ws = ExcelWorkbook[sheet_name]
                thin_border = Border(left=Side(style='thin'),
                                     right=Side(style='thin'),
                                     top=Side(style='thin'),
                                     bottom=Side(style='thin'))
                max_column = ws.max_column + 1
                max_row = ws.max_row + 1
                for col in range(1, max_column):
                    ws.cell(1, col).fill = PatternFill("solid", start_color="d9e1f2")
                    for row in range(1, max_row):
                        ws.cell(row, col).border = thin_border
                        ws.cell(row, col).font = Font(name='Trebuchet MS')
                        if row == max_row - 1 and sheet_name == 'Financial Simulation':
                            ws.cell(row, col).fill = PatternFill("solid", start_color="C5D9F1")

            ExcelWorkbook.save(fileName)
            os.startfile(os.path.dirname(fileName))
        except:
            pass


    # inventory optimizer
    def opt_nonzerostd(self, row):
        row = self.opt_filtered_row(row)
        row = row.replace({0: pd.NA, '0': pd.NA, '': pd.NA, 'None': pd.NA})
        row = pd.Series(row.dropna().reset_index(drop=True))

        return row.std(ddof=0)

    # inventory optimizer
    def opt_filtered_row(self, row):
        first_index = row.replace({0: pd.NA, '0': pd.NA, '': pd.NA, 'None': pd.NA}).first_valid_index()
        row = row[first_index:]
        row = row.fillna(0)
        return pd.Series(row.replace({pd.NA: 0}))

    # inventory optimizer
    def opt_length(self, row):
        return len(self.opt_filtered_row(row))

    # inventory optimizer (main)
    def opt_process_file(self):
        if self.df is None: return
        df = self.df
        df = df.dropna(how='all')
        df = df.fillna(0)
        df = df.replace({0: pd.NA, '0': pd.NA, '': pd.NA, 'None': pd.NA})

        start_col = df.columns.get_loc('Description')

        self.end_col = end_col = len(df.columns) - 3
        current_min_col = df.iloc[:, self.end_col + 1]
        current_max_col = df.iloc[:, self.end_col + 2]

        LT_Adjustment = 0 if self.ui.LT_Adjustment.text() in [None, 'None', '', ' '] else int(
            self.ui.LT_Adjustment.text())
        Alpha = 0 if self.ui.Alpha.text() in [None, 'None', '', ' '] else float(self.ui.Alpha.text())
        ReviewPeriod = 0 if self.ui.ReviewPeriod.text() in [None, 'None', '', ' '] else float(
            self.ui.ReviewPeriod.text())
        ReviewPeriod = ReviewPeriod / 30.41666667
        df['Total'] = df.iloc[:, start_col + 1:end_col].apply(self.opt_filtered_row, axis=1).sum(axis=1)
        df['LifeTime'] = df.iloc[:, start_col + 1:end_col].apply(self.opt_length, axis=1)
        df['Min Order'] = df.iloc[:, start_col + 1:end_col].apply(self.opt_filtered_row, axis=1).replace(
            {'0': pd.NA, 0: pd.NA}).min(axis=1)
        df['Max Order'] = df.iloc[:, start_col + 1:end_col].apply(self.opt_filtered_row, axis=1).max(axis=1)
        df['Adjusted LT (days)'] = df.iloc[:, end_col] + LT_Adjustment
        df['Adjusted LT (month)'] = df['Adjusted LT (days)'] / 30.41666667
        df['Average Demand'] = df['Total'] / df[
            'LifeTime']  # df.iloc[:, start_col + 1:end_col].apply(self.get_filtered_row,axis=1).mean(axis=1) #df.iloc[:, pd.Series.first_valid_index:end_col].mean(axis=1)
        df['Demand over LT'] = df['Average Demand'] * df['Adjusted LT (month)']

        df['Std. Deviation of Demand'] = df.iloc[:, start_col + 1:end_col].apply(self.opt_nonzerostd,
                                                                                 1)  # .std(axis=1).replace({0:0,
        # pd.NA:0,np.nan:0})

        df['Std. Dev. of Demand over LT'] = df['Std. Deviation of Demand'] * np.sqrt(
            df['Adjusted LT (month)'].astype(float))
        # std_of_demand_over_LT =
        df['Safety Stock'] = scipy.stats.norm.ppf(Alpha, 0, 1) * df[
            'Std. Dev. of Demand over LT']  # norm.ppf(alpha-Alpha,mu-,sigma-sd) norm.inv(0.98,0,1)
        df['Min'] = scipy.stats.norm.ppf(Alpha, 0, 1) * df['Std. Dev. of Demand over LT'] + df['Demand over LT']
        # 'No BenchMark','Current Min','Max Order','Both','Slow Moving Items'
        if str(self.ui.comboBox.currentText()) == 'No BenchMark':
            df['Adjusted MIN'] = df['Min']
        elif str(self.ui.comboBox.currentText()) == 'Current Min':
            df['Adjusted MIN'] = np.where(df['Min'] < current_min_col, current_min_col, df['Min'])
        elif str(self.ui.comboBox.currentText()) == 'Max Order':
            df['Adjusted MIN'] = np.where(df['Min'] < df['Max Order'], df['Max Order'], df['Min'])
        elif str(self.ui.comboBox.currentText()) == 'Both':
            max_result = np.where(current_min_col < df['Max Order'], df['Max Order'], current_min_col)
            df['Adjusted MIN'] = np.where(df['Min'] < max_result, max_result, df['Min'])
        elif str(self.ui.comboBox.currentText()) == 'Slow Moving Items':
            df['Adjusted MIN'] = np.ceil(
                np.where(df['Demand over LT'] > 10, df['Demand over LT'] + df['Average Demand'],
                         df['Demand over LT'] + 1)).astype(int)
            df['MAX'] = np.ceil(df['Adjusted MIN'] + df['Demand over LT']).astype(int)

        elif str(self.ui.comboBox.currentText()) == 'Compare All':
            df['Adjusted MIN-No BenchMark'] = df['Min']
            df['Adjusted MIN-Current Min'] = np.where(df['Min'] < current_min_col, current_min_col, df['Min'])
            df['Adjusted MIN-Max Order'] = np.where(df['Min'] < df['Max Order'], df['Max Order'], df['Min'])
            max_result = np.where(current_min_col < df['Max Order'], df['Max Order'], current_min_col)
            df['Adjusted MIN-Both'] = np.where(df['Min'] < max_result, max_result, df['Min'])
            df['Adjusted MIN-Slow Moving Items'] = np.ceil(
                np.where(df['Demand over LT'] > 10, df['Demand over LT'] + df['Average Demand'],
                         df['Demand over LT'] + 1)).astype(int)

            df['MAX-No BenchMark'] = (df['Average Demand'] * ReviewPeriod) + df['Adjusted MIN-No BenchMark']
            df['MAX-Current Min'] = (df['Average Demand'] * ReviewPeriod) + df['Adjusted MIN-Current Min']
            df['MAX-Max Order'] = (df['Average Demand'] * ReviewPeriod) + df['Adjusted MIN-Max Order']
            df['MAX-Both'] = (df['Average Demand'] * ReviewPeriod) + df['Adjusted MIN-Both']
            df['MAX-Slow Moving Items'] = np.ceil(df['Adjusted MIN-Slow Moving Items'] + df['Demand over LT']).astype(
                int)
        # need addition over here
        if str(self.ui.comboBox.currentText()) != 'Slow Moving Items' and str(
                self.ui.comboBox.currentText()) != 'Compare All':
            df['MAX'] = (df['Average Demand'] * ReviewPeriod) + df['Adjusted MIN']

        # std_of_demand=df['Std. Deviation of Demand'].replace({0:0.0001})
        # std_of_demand_over_LT=std_of_demand*np.sqrt(df['Adjusted LT (month)'].astype(float))
        # df['Adjusted STD DEV of Demand over LT']=df['Std. Dev. of Demand over LT'].replace({0: 0.001})
        if str(self.ui.comboBox.currentText()) == 'Compare All':
            df['Cycle Service Level-No BenchMark'] = scipy.stats.norm.cdf(
                (df['Adjusted MIN-No BenchMark'].astype(float) + 0.0056).round(decimals=4),
                df['Demand over LT'].astype(float).round(decimals=3),
                df['Std. Dev. of Demand over LT'].replace({0: 0.001}).astype(
                    float).round(decimals=4))

            df['Cycle Service Level-Current Min'] = scipy.stats.norm.cdf(
                (df['Adjusted MIN-Current Min'].astype(float) + 0.0056).round(decimals=4),
                df['Demand over LT'].astype(float).round(decimals=3),
                df['Std. Dev. of Demand over LT'].replace({0: 0.001}).astype(
                    float).round(decimals=4))

            df['Cycle Service Level-Max Order'] = scipy.stats.norm.cdf(
                (df['Adjusted MIN-Max Order'].astype(float) + 0.0056).round(decimals=4),
                df['Demand over LT'].astype(float).round(decimals=3),
                df['Std. Dev. of Demand over LT'].replace({0: 0.001}).astype(
                    float).round(decimals=4))

            df['Cycle Service Level-Both'] = scipy.stats.norm.cdf(
                (df['Adjusted MIN-Both'].astype(float) + 0.0056).round(decimals=4),
                df['Demand over LT'].astype(float).round(decimals=3),
                df['Std. Dev. of Demand over LT'].replace({0: 0.001}).astype(
                    float).round(decimals=4))

            df['Cycle Service Level-Slow Moving Items'] = scipy.stats.norm.cdf(
                (df['Adjusted MIN-Slow Moving Items'].astype(float) + 0.0056).round(decimals=4),
                df['Demand over LT'].astype(float).round(decimals=3),
                df['Std. Dev. of Demand over LT'].replace({0: 0.001}).astype(
                    float).round(decimals=4))
        else:
            df['Cycle Service Level'] = scipy.stats.norm.cdf(
                (df['Adjusted MIN'].astype(float) + 0.0056).round(decimals=4),
                df['Demand over LT'].astype(float).round(decimals=3),
                df['Std. Dev. of Demand over LT'].replace({0: 0.001}).astype(
                    float).round(decimals=4))
        # NormalDist(mu=0, sigma=1).cdf(1.96)

        # Formatiing
        df['Adjusted LT (month)'] = df['Adjusted LT (month)'].astype(float).round(1)
        df['Average Demand'] = df['Average Demand'].astype(float).round(1)
        df['Demand over LT'] = df['Demand over LT'].astype(float).round(1)
        df['Std. Deviation of Demand'] = df['Std. Deviation of Demand'].astype(float).round(1)
        df['Std. Dev. of Demand over LT'] = df['Std. Dev. of Demand over LT'].astype(float).round(1)
        df['Safety Stock'] = df['Safety Stock'].astype(float).round(1)
        df['Min'] = df['Min'].astype(float).round(1)
        if str(self.ui.comboBox.currentText()) == 'Compare All':
            for item in ['No BenchMark', 'Current Min', 'Max Order', 'Both', 'Slow Moving Items']:
                df['Adjusted MIN-{}'.format(item)] = df['Adjusted MIN-{}'.format(item)].astype(float).round(1)
                df['MAX-{}'.format(item)] = df['MAX-{}'.format(item)].astype(float).round(1)
                df['Cycle Service Level-{}'.format(item)] = df['Cycle Service Level-{}'.format(item)].astype(
                    float).round(5)
        else:
            df['Adjusted MIN'] = df['Adjusted MIN'].astype(float).round(1)
            df['MAX'] = df['MAX'].astype(float).round(1)
            df['Cycle Service Level'] = df['Cycle Service Level'].astype(float).round(5)

        self.df = df

        print('Output generated successfully')
        self.opt_save_output()

    # inventory optimizer
    def opt_save_output(self):
        fileName = 'output/opt_output.xlsx'
        if fileName is None: return
        try:
            self.df.to_excel(fileName, index=False)

            ExcelWorkbook = load_workbook(fileName)
            thin_border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))
            ws = ExcelWorkbook.active
            max_column = ws.max_column + 1
            max_row = ws.max_row + 1
            for col in range(1, max_column):
                ws.cell(1, col).fill = PatternFill("solid", start_color="d9e1f2")
                for row in range(1, max_row):
                    ws.cell(row, col).border = thin_border

            ExcelWorkbook.save(fileName)
        except:
            print('Unable To Save File')
            return

        self.df = None

    # simulation
    def sim_process_row(self, row, ci, simulation_duration_type):
        # to remove the data which do not have itemNumber
        if pd.isna(row['ItemNumber']):
            return

        # current stock is updated on every column
        if ci == 1:
            cs = row['Current Stock']
            prev_stock = 0 if pd.isna(cs) else cs
        else:
            prev_stock = row[simulation_duration_type + str(ci - 1)]

        # current column index
        current_index = simulation_duration_type + str(ci)

        # difference between prev_stock and average monthly consumption
        deviation = self.ui.CTS_Consumption_Dev.value()
        if simulation_duration_type == 'M':
            if float(deviation) == float(0):
                average_consumption = 0.00 if pd.isna(row['Avrg. Consumption/Mnth']) else float(
                    row['Avrg. Consumption/Mnth'])
            else:
                average_consumption = 0.00 if pd.isna(row['Avrg. Consumption/Mnth']) else float(
                    row['Avrg. Consumption/Mnth'])
                limit = float(average_consumption * (deviation / 100))
                average_consumption = random.uniform(average_consumption - limit, average_consumption + limit)
        elif simulation_duration_type == 'D':
            if float(deviation) == float(0):
                average_consumption = 0.00 if pd.isna(row['Avrg. Consumption/Mnth']) else float(
                    row['Avrg. Consumption/Mnth'])
                average_consumption = average_consumption / 30.416
            else:
                average_consumption = 0.00 if pd.isna(row['Avrg. Consumption/Mnth']) else float(
                    row['Avrg. Consumption/Mnth'])

                average_consumption = average_consumption / 30.416
                limit = float(average_consumption * (deviation / 100))
                average_consumption = random.uniform(average_consumption - limit, average_consumption + limit)

        diff = prev_stock - average_consumption

        # if no order already placed
        if pd.isna(row['Prev_Order']):
            row[current_index] = diff

            if row[current_index] <= row['MIN (Reorder Point)']:
                # index + diff
                order = row['MAX (Order Up To)'] - diff
                if simulation_duration_type == 'M':

                    if row['Rounded Lead Time'] < 1:
                        # execute order in same month
                        row[current_index] = diff + order
                        row['Prev_Order'] = pd.NA

                        # check if we need to check again or not
                    else:

                        row['Prev_Order'] = str(int(current_index.lstrip(simulation_duration_type)) + math.floor(
                            row['Rounded Lead Time'])) + '+' + str(order)
                else:
                    row['Prev_Order'] = str(
                        int(current_index.lstrip(simulation_duration_type)) + row['Rounded Lead Time']) + '+' + str(
                        order)
        else:
            if str(int(current_index.lstrip(simulation_duration_type))) == row['Prev_Order'].split('+')[0].rstrip('.0'):
                row[current_index] = diff + float(row['Prev_Order'].split('+')[1])
                row['Prev_Order'] = pd.NA

                if row[current_index] <= row['MIN (Reorder Point)']:
                    # index + diff
                    order = row['MAX (Order Up To)'] - row[current_index]
                    row['Prev_Order'] = str(
                        int(current_index.lstrip(simulation_duration_type)) + row['Rounded Lead Time']) + '+' + str(
                        order)
            else:
                row[current_index] = diff

        return row

    # simulation (main)
    def sim_process_file(self):
        if self.df is None: return
        df = self.df

        simulation_time = self.ui.Simulation_period.value()

        if simulation_time == 0:
            return

        simulation_time_type = str(self.ui.sumulation_duration_type.currentText()).lower().strip()

        default_price = self.ui.price.value()

        # df = pd.read_excel('Book5.xlsx')
        df = df.dropna(how='all')
        df = df.fillna(0)
        df = df.replace({0: pd.NA, '0': pd.NA, '': pd.NA, 'None': pd.NA})

        # ...................................

        if simulation_time_type == 'months':
            df['Lead Time (months)'] = df['Lead Time (days)'].astype(float) / 30.416
            if self.ui.Round_LT.isChecked():
                df['Rounded Lead Time'] = df['Lead Time (months)'].replace({pd.NA: 0}).round().astype(int)
            else:
                df['Rounded Lead Time'] = df['Lead Time (months)'].replace({pd.NA: 0}).astype(float)
            simulation_duration_type = 'M'
        elif simulation_time_type == 'days':
            simulation_duration_type = 'D'
            df['Rounded Lead Time'] = df['Lead Time (days)'].replace({pd.NA: 0}).astype(float)

        df['Prev_Order'] = pd.NA

        for rep in range(simulation_time):
            df['{}{}'.format(simulation_duration_type, rep + 1)] = ''
            new_data = df.apply(self.sim_process_row, axis=1, ci=rep + 1,
                                simulation_duration_type=simulation_duration_type)
            # if not pd.isna(new_data['Prev_Order']):
            df['Prev_Order'] = new_data['Prev_Order']
            df['{}{}'.format(simulation_duration_type, rep + 1)] = new_data[
                '{}{}'.format(simulation_duration_type, rep + 1)]

        df.drop(columns=['Prev_Order'], inplace=True)

        # calculate second sheet
        fs_df = df[['ItemNumber', 'Description', 'Price/Unit']].copy()

        for column in range(simulation_time):
            colm_title = '{}{}'.format(simulation_duration_type, column + 1)
            price_column = df['Price/Unit']
            price_column.replace({pd.NA: default_price}, inplace=True)
            amount_column = df[colm_title] * price_column
            fs_df[colm_title] = amount_column

        def remove_neg(x):
            if 'M' in x.name:
                return x.clip(lower=0)
            else:
                return x

        if self.ui.Remove_Negs.isChecked():
            fs_df = fs_df.apply(remove_neg)

        fs_df.loc['Total'] = fs_df.sum(axis=0)
        fs_df.iloc[-1, 1] = ''
        fs_df.iloc[-1, 0] = 'Total'

        # ...............

        # ...................................

        self.df = df
        self.fs_df = fs_df

        self.custom_dialog.label_4.setText('    File Successfully Created !')
        self.custom_dialog.Export.setText('Export')
        try:
            self.custom_dialog.Export.disconnect()
        except:
            pass
        self.custom_dialog.show()
        self.custom_dialog.Export.clicked.connect(self.sim_save_output)
        try:
            self.ui.Process_1.disconnect()
        except:
            pass
        self.ui.Process_1.clicked.connect(lambda: self.upload_df(self.ExcelPath_1, self.ui.Process_1))

    # sumulation
    def sim_save_output(self):
        fileName='output/sim_output.xlsx'
        if fileName is None: return
        try:
            writer = pd.ExcelWriter(fileName, engine='openpyxl', mode='w+')

            self.df.to_excel(writer, sheet_name='Inventory Simulation', index=False)
            self.fs_df.to_excel(writer, sheet_name='Financial Simulation', index=False)

            # writer.save()
            writer.close()

            ExcelWorkbook = load_workbook(fileName)
            print(ExcelWorkbook.sheetnames)
            # ExcelWorkbook.remove(ExcelWorkbook['Sheet'])
            for sheet_name in ['Inventory Simulation', 'Financial Simulation']:
                ws = ExcelWorkbook[sheet_name]
                thin_border = Border(left=Side(style='thin'),
                                     right=Side(style='thin'),
                                     top=Side(style='thin'),
                                     bottom=Side(style='thin'))
                max_column = ws.max_column + 1
                max_row = ws.max_row + 1
                for col in range(1, max_column):
                    ws.cell(1, col).fill = PatternFill("solid", start_color="d9e1f2")
                    for row in range(1, max_row):
                        ws.cell(row, col).border = thin_border
                        ws.cell(row, col).font = Font(name='Trebuchet MS')
                        if row == max_row - 1 and sheet_name == 'Financial Simulation':
                            ws.cell(row, col).fill = PatternFill("solid", start_color="C5D9F1")

            ExcelWorkbook.save(fileName)

        except Exception as e:
            print(e)
            print('Unable To Save File')
            return
